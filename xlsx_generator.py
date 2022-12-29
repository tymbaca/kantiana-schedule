import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.worksheet.dimensions import Dimension
from openpyxl.utils import get_column_letter, get_column_interval

from classes import *
from settings import *


SIDE_PERIODS_COLUMN = "A"

GROUP_COLUMNS = {
    1: "B",
    2: "D",
    3: "F"
}

LESSON_PERIODS = {
    "08:30 - 10:00": 0,
    "10:10 - 11:40": 1,
    "11:50 - 13:20": 2,
    "13:50 - 15:20": 3,
    "15:30 - 17:00": 4,
    "17:10 - 18:40": 5
}

WEEKDAY_DICT = {
    0: "Понедельник",
    1: "Вторник",
    2: "Среда",
    3: "Четверг",
    4: "Пятница",
    5: "Суббота",
    6: "Воскресенье"
}

ROWS_PER_DAY = len(LESSON_PERIODS) + 1  # Because of daily header

# Alignments
LESSONS_ALIGNMENT = Alignment(horizontal="center", vertical="top", wrap_text=True)
DOUBLE_LESSONS_ALIGNMENT = Alignment(horizontal="center", vertical="top", wrap_text=True)
SIDE_PERIODS_ALIGNMENT = Alignment(vertical="center", horizontal="right")
DAILY_HEADERS_ALIGNMENT = Alignment(horizontal="center", vertical="bottom")
MAIN_HEADERS_ALIGNMENT = Alignment(horizontal="center", vertical="bottom")

if IS_LESSONS_UNITY_ALIGNMENT:
    DOUBLE_LESSONS_ALIGNMENT = LESSONS_ALIGNMENT

# Fonts
LESSONS_FONT = Font()
DOUBLE_LESSONS_FONT = Font()
SIDE_PERIODS_FONT = Font(bold=True)
DAILY_HEADERS_FONT = Font(bold=True, size=24)
MAIN_HEADERS_FONT = Font(bold=True, size=16)

if IS_LESSONS_UNITY_FONT:
    DOUBLE_LESSONS_FONT = LESSONS_FONT


LESSONS_HEIGHT = 95
DOUBLE_LESSONS_HEIGHT = 125
LESSONS_WIDTH = 22


def _fill_side_periods(ws: Worksheet, origin_row: int):
    for period, index in LESSON_PERIODS.items():
        cell_index = f"{SIDE_PERIODS_COLUMN}{1 + origin_row + index}"
        ws[cell_index] = period
        ws[cell_index].font = SIDE_PERIODS_FONT
        ws[cell_index].alignment = SIDE_PERIODS_ALIGNMENT

    ws.column_dimensions[SIDE_PERIODS_COLUMN].width = 12


def _next_column(char) -> str:
    return chr(ord(char) + 1)


def _get_columns(group: Schedule.group) -> tuple[str, str]:
    left_column = GROUP_COLUMNS[group]
    right_column = _next_column(left_column)
    return left_column, right_column


def _get_row(time: Lesson.time, origin_row) -> int:
    row_offset = LESSON_PERIODS[time]
    row = 1 + origin_row + row_offset
    return row


def _fill_lesson_cell(ws: Worksheet, lesson: Lesson, column: str, row: int, is_double_lesson: bool = False):
    cell_index = f"{column}{row}"

    # Filling content
    if not is_double_lesson:
        cell_content = f"({lesson.type_})\n{lesson.subject}\n{lesson.teacher}\n{lesson.location}\n{lesson.subgroup}"
    else:
        cell_content = f"({lesson.type_})\n{lesson.subject}\n{lesson.teacher}\n{lesson.location}\n{lesson.subgroup}"

    ws[cell_index] = cell_content

    # Selecting lesson formatting type
    if not is_double_lesson:
        cell_font = LESSONS_FONT
        cell_alignment = LESSONS_ALIGNMENT
    else:
        cell_font = DOUBLE_LESSONS_FONT
        cell_alignment = DOUBLE_LESSONS_ALIGNMENT

    # Formatting
    ws[cell_index].font = cell_font
    ws[cell_index].alignment = cell_alignment

    # Setting column/row width/height
    ws.column_dimensions[column].width = LESSONS_WIDTH
    if is_double_lesson:
        ws.row_dimensions[row].height = DOUBLE_LESSONS_HEIGHT
    else:
        ws.row_dimensions[row].height = LESSONS_HEIGHT


def _fill_double_lesson(ws: Worksheet, schedule_item: ScheduleItem, left_col, right_col, row: int):
    for lesson in schedule_item.period:
        if lesson.subgroup == "1 подгруппа":
            _fill_lesson_cell(ws, lesson, left_col, row, is_double_lesson=True)
        elif lesson.subgroup == "2 подгруппа":
            _fill_lesson_cell(ws, lesson, right_col, row, is_double_lesson=True)
        else:
            raise Exception(f"Опять какая то хуйня. В двойном ScheduleItem встретилась подгруппа {lesson.subgroup}")


def _fill_single_lesson(ws: Worksheet, schedule_item: ScheduleItem, left_col, right_col, row: int):
    ws.merge_cells(f"{left_col}{row}:{right_col}{row}")

    lesson: Lesson = schedule_item.period
    _fill_lesson_cell(ws, lesson, left_col, row)


def _fill_lessons(ws: Worksheet, schedule: Schedule, origin_row: int):
    if schedule.items is None:
        return None

    for schedule_item in schedule.items:
        # Finding columns and row indexes
        left_column, right_column = _get_columns(schedule.group)
        if schedule_item.item_type == ScheduleItemType.SINGLE:
            row: int = _get_row(schedule_item.period.time, origin_row)
        else:
            row: int = _get_row(schedule_item.period[0].time, origin_row)

        # Setting column/row width/height
        ws.column_dimensions[left_column].width = 22
        ws.column_dimensions[right_column].width = 22
        ws.row_dimensions[row].height = 64

        if schedule_item.item_type == ScheduleItemType.SINGLE:
            _fill_single_lesson(ws, schedule_item, left_column, right_column, row)
        elif schedule_item.item_type == ScheduleItemType.DOUBLE:
            _fill_double_lesson(ws, schedule_item, left_column, right_column, row)
        else:
            raise Exception("Опять какая-то поебота, которая никогда не случится. Проверь ScheduleItemType!")


def _fill_daily_header(ws: Worksheet, schedule: Schedule, row_number: int):
    date = schedule.date
    weekday = WEEKDAY_DICT[date.weekday()]

    header_content = f"{date.strftime(DATE_FORMAT)} {weekday}"
    ws.merge_cells(f"B{row_number}:G{row_number}")
    ws[f"B{row_number}"] = header_content

    # Formatting
    ws[f"B{row_number}"].font = DAILY_HEADERS_FONT
    ws[f"B{row_number}"].alignment = DAILY_HEADERS_ALIGNMENT
    ws.row_dimensions[int(row_number)].height = 30


def _fill_daily_schedule(ws, schedule, day_count):
    origin_row = day_count * ROWS_PER_DAY + 2

    _fill_daily_header(ws, schedule, origin_row)
    _fill_lessons(ws, schedule, origin_row)
    _fill_side_periods(ws, origin_row)


def _fill_contents(ws: Worksheet, schedules: list[Schedule]):
    day_count = 0
    for schedule in schedules:
        _fill_daily_schedule(ws, schedule, day_count)
        day_count += 1


def _fill_headers(ws: Worksheet):
    ws.merge_cells("B1:C1")
    ws.merge_cells("D1:E1")
    ws.merge_cells("F1:G1")

    time_header_id = "A1"
    ws[time_header_id] = "Время"
    ws[time_header_id].font = MAIN_HEADERS_FONT
    ws[time_header_id].alignment = MAIN_HEADERS_ALIGNMENT

    for group, column in GROUP_COLUMNS.items():
        cell_index = f"{column}1"
        ws[cell_index] = f"{group} Группа"
        ws[cell_index].font = MAIN_HEADERS_FONT
        ws[cell_index].alignment = MAIN_HEADERS_ALIGNMENT


def _fill_sheet(wb: openpyxl.Workbook, groups_schedule_lists: list[list[Schedule]]):
    ws: Worksheet = wb.active
    _fill_headers(ws)
    for schedule_list in groups_schedule_lists:
        _fill_contents(ws, schedule_list)


def generate(groups_schedule_lists: list[list[Schedule]], filename: str):
    wb = openpyxl.Workbook()
    _fill_sheet(wb, groups_schedule_lists)
    wb.save(filename)


if __name__ == "__main__":
    generate([], "bububu.xlsx")
